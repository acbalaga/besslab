import sys
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app import _build_hourly_summary_workbook
from services.simulation_core import HourlyLog


def test_hourly_summary_workbook_contains_year_sheets():
    hours = 24
    timestamp = pd.date_range("2024-01-01", periods=hours, freq="h")
    logs = HourlyLog(
        hod=np.arange(hours),
        pv_mw=np.ones(hours),
        pv_to_contract_mw=np.ones(hours),
        bess_to_contract_mw=np.zeros(hours),
        charge_mw=np.zeros(hours),
        discharge_mw=np.zeros(hours),
        soc_mwh=np.linspace(0.0, 10.0, hours),
        timestamp=timestamp.to_numpy(),
    )
    workbook_bytes = _build_hourly_summary_workbook({1: logs, 2: logs})
    workbook = load_workbook(BytesIO(workbook_bytes))
    assert workbook.sheetnames == ["Year 1", "Year 2"]
